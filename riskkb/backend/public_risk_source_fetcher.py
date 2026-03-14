#!/usr/bin/env python3
"""Fetch public non-PubMed risk-taxonomy sources."""

from __future__ import annotations

import argparse
import json
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
KNOWLEDGE_DIR = ROOT / "knowledge"
DEFAULT_OUTPUT_DIR = KNOWLEDGE_DIR / "derived" / "risk_taxonomy_raw_public_sources"

USER_AGENT = "standalone-food-risk-kb/1.0"
FOOD_TERMS = [
    "food",
    "milk",
    "dairy",
    "cheese",
    "yogurt",
    "yoghurt",
    "butter",
    "cream",
    "whey",
    "infant formula",
    "formula",
    "egg",
    "meat",
    "poultry",
    "chicken",
    "beef",
    "pork",
    "fish",
    "seafood",
    "shellfish",
    "shrimp",
    "oyster",
    "vegetable",
    "fruit",
    "juice",
    "rice",
    "grain",
    "cereal",
    "salad",
    "sprout",
    "nut",
    "peanut",
]
HAZARD_TERMS = [
    "aflatoxin",
    "arsenic",
    "cadmium",
    "campylobacter",
    "chloramphenicol",
    "clostridium",
    "cronobacter",
    "deoxynivalenol",
    "dioxin",
    "lead",
    "listeria",
    "melamine",
    "mercury",
    "nitrofuran",
    "nitrite",
    "nitrate",
    "ochratoxin",
    "patulin",
    "pcb",
    "pfas",
    "salmonella",
    "staphylococcus",
    "vibrio",
    "zearalenone",
]
FDA_FOOD_OFFICE_TERMS = [
    "human foods program",
    "food safety and applied nutrition",
    "imports",
    "human and animal food operations",
    "veterinary medicine",
]
FDA_FOOD_SUBJECT_TERMS = [
    "food",
    "fsvp",
    "seafood haccp",
    "shell egg",
    "low acid canned food",
    "acidified",
    "produce",
    "juice haccp",
    "infant formula",
    "listeria",
    "salmonella",
    "allergen",
    "insanitary",
]


def _json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, sort_keys=True)


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_json_dumps(data) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(_json_dumps(row) + "\n")


def _request_bytes(url: str, timeout: int = 60) -> bytes:
    last_error: Exception | None = None
    for attempt in range(4):
        request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return response.read()
        except (urllib.error.URLError, TimeoutError) as exc:
            last_error = exc
            time.sleep(1.0 + 1.5 * attempt)
    if last_error is not None:
        raise last_error
    raise RuntimeError("request failed without exception")


def _request_text(url: str, timeout: int = 60) -> str:
    return _request_bytes(url, timeout=timeout).decode("utf-8", "ignore")


def _request_json(url: str, timeout: int = 60) -> Any:
    return json.loads(_request_text(url, timeout=timeout))


def _slugify(text: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return value or "column"


def _looks_food_related(text: str) -> bool:
    lowered = text.lower()
    return any(term in lowered for term in FOOD_TERMS)


def _looks_fda_food_letter(item: dict[str, Any]) -> bool:
    office = str(item.get("issuing_office", "") or "").lower()
    subject = str(item.get("subject", "") or "").lower()
    company = str(item.get("company_name", "") or "").lower()
    if any(term in office for term in FDA_FOOD_OFFICE_TERMS):
        return True
    text = " ".join([office, subject, company])
    return any(term in text for term in FDA_FOOD_SUBJECT_TERMS)


class PublicRiskSourceFetcher:
    def __init__(self, output_dir: Path = DEFAULT_OUTPUT_DIR) -> None:
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def fetch_all(self) -> dict[str, Any]:
        manifest = {
            "generated_at": datetime.now(UTC).isoformat(),
            "output_dir": str(self.output_dir),
        }
        manifest["cdc_nors_food_records"] = self.fetch_cdc_nors_food()
        manifest["fda_warning_food_records"] = self.fetch_fda_warning_letters()
        manifest["fda_outbreak_report_records"] = self.fetch_fda_outbreak_reports()
        manifest["jecfa_records"] = self.fetch_jecfa_records()
        manifest["openfoodtox_assets"] = self.fetch_openfoodtox_index()
        _write_json(self.output_dir / "public_sources_manifest.json", manifest)
        return manifest

    def fetch_cdc_nors_food(self) -> int:
        output_dir = self.output_dir / "cdc_nors"
        output_dir.mkdir(parents=True, exist_ok=True)
        rows: list[dict[str, Any]] = []
        limit = 50000
        offset = 0

        while True:
            params = urllib.parse.urlencode(
                {
                    "$select": "*",
                    "$where": "primary_mode='Food'",
                    "$limit": str(limit),
                    "$offset": str(offset),
                }
            )
            payload = _request_json(f"https://data.cdc.gov/resource/5xkq-dg7x.json?{params}")
            if not payload:
                break
            for index, item in enumerate(payload):
                text = " ".join(
                    [
                        item.get("etiology", ""),
                        item.get("food_vehicle", ""),
                        item.get("food_contaminated_ingredient", ""),
                        item.get("ifsac_category", ""),
                        item.get("location_of_preparation", ""),
                    ]
                )
                rows.append(
                    {
                        "record_type": "event_evidence",
                        "source": "cdc_nors",
                        "source_url_or_id": item.get("cdc_date_case_id")
                        or f"nors_{offset + index}",
                        "authority": "CDC",
                        "evidence_type": "outbreak",
                        "year": item.get("year", ""),
                        "month": item.get("month", ""),
                        "state": item.get("state", ""),
                        "etiology": item.get("etiology", ""),
                        "etiology_status": item.get("etiology_status", ""),
                        "food_vehicle": item.get("food_vehicle", ""),
                        "food_contaminated_ingredient": item.get("food_contaminated_ingredient", ""),
                        "ifsac_category": item.get("ifsac_category", ""),
                        "setting": item.get("setting", ""),
                        "illnesses": item.get("illnesses", ""),
                        "hospitalizations": item.get("hospitalizations", ""),
                        "deaths": item.get("deaths", ""),
                        "suspected_food_related": _looks_food_related(text),
                        "raw": item,
                    }
                )
            offset += len(payload)
            print(_json_dumps({"source": "cdc_nors", "fetched": offset}), flush=True)
            if len(payload) < limit:
                break
            time.sleep(0.2)

        _write_jsonl(output_dir / "cdc_nors_foodborne.jsonl", rows)
        _write_json(
            output_dir / "cdc_nors_foodborne_manifest.json",
            {
                "generated_at": datetime.now(UTC).isoformat(),
                "records": len(rows),
                "source_url": "https://data.cdc.gov/resource/5xkq-dg7x.json",
            },
        )
        return len(rows)

    def fetch_fda_warning_letters(self) -> int:
        output_dir = self.output_dir / "fda_warning_letters"
        output_dir.mkdir(parents=True, exist_ok=True)
        xlsx_url = (
            "https://www.fda.gov/inspections-compliance-enforcement-and-criminal-investigations/"
            "compliance-actions-and-activities/warning-letters/datatables-data?page&_format=xlsx"
        )
        xlsx_path = output_dir / "warning-letters.xlsx"
        xlsx_path.write_bytes(_request_bytes(xlsx_url))

        workbook = load_workbook(xlsx_path, read_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [_slugify(str(cell or "")) for cell in next(rows_iter)]
        all_rows: list[dict[str, Any]] = []
        food_rows: list[dict[str, Any]] = []

        for index, row in enumerate(rows_iter, 1):
            item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            record = {
                "record_type": "event_evidence",
                "source": "fda_warning_letters",
                "source_url_or_id": item.get("warning_letter_url") or item.get("title") or f"warning_{index}",
                "authority": "FDA",
                "evidence_type": "warning_letter",
                "is_food_related": _looks_fda_food_letter(item),
                "raw": item,
            }
            all_rows.append(record)
            if record["is_food_related"]:
                food_rows.append(record)

        _write_jsonl(output_dir / "warning_letters_all.jsonl", all_rows)
        _write_jsonl(output_dir / "warning_letters_food_related.jsonl", food_rows)
        _write_json(
            output_dir / "warning_letters_manifest.json",
            {
                "generated_at": datetime.now(UTC).isoformat(),
                "all_records": len(all_rows),
                "food_related_records": len(food_rows),
                "xlsx_url": xlsx_url,
            },
        )
        return len(food_rows)

    def fetch_fda_outbreak_reports(self) -> int:
        output_dir = self.output_dir / "fda_outbreaks"
        output_dir.mkdir(parents=True, exist_ok=True)
        reports_url = "https://www.fda.gov/food/outbreaks-foodborne-illness/outbreak-investigation-reports"
        core_url = "https://www.fda.gov/food/hfp-constituent-updates/core-outbreak-investigation-table-issued"

        reports_html = _request_text(reports_url)
        core_html = _request_text(core_url)
        soup = BeautifulSoup(reports_html, "html.parser")

        links: list[str] = []
        for anchor in soup.find_all("a", href=True):
            href = anchor["href"]
            if "/food/outbreaks-foodborne-illness/" not in href:
                continue
            if "outbreak-investigation-reports" in href:
                continue
            if href.startswith("/"):
                href = f"https://www.fda.gov{href}"
            links.append(href)
        report_links = list(dict.fromkeys(links))

        records: list[dict[str, Any]] = []
        for index, link in enumerate(report_links, 1):
            html = _request_text(link)
            page = BeautifulSoup(html, "html.parser")
            title = ""
            if page.title and page.title.string:
                title = page.title.string.strip()
            description = ""
            desc = page.find("meta", attrs={"name": "description"})
            if desc:
                description = (desc.get("content") or "").strip()
            body_text = " ".join(page.get_text(" ", strip=True).split())[:5000]
            records.append(
                {
                    "record_type": "event_evidence",
                    "source": "fda_core_outbreak",
                    "source_url_or_id": link,
                    "authority": "FDA",
                    "evidence_type": "outbreak_report",
                    "title": title,
                    "description": description,
                    "is_food_related": _looks_food_related(f"{title} {description} {body_text}"),
                    "body_excerpt": body_text,
                }
            )
            print(_json_dumps({"source": "fda_outbreak_report", "done": index, "total": len(report_links)}), flush=True)
            time.sleep(0.2)

        core_page = BeautifulSoup(core_html, "html.parser")
        core_text = " ".join(core_page.get_text(" ", strip=True).split())
        core_meta = {
            "record_type": "event_evidence",
            "source": "fda_core_outbreak",
            "source_url_or_id": core_url,
            "authority": "FDA",
            "evidence_type": "outbreak_table",
            "title": core_page.title.string.strip() if core_page.title and core_page.title.string else "",
            "body_excerpt": core_text[:8000],
        }

        _write_jsonl(output_dir / "fda_outbreak_reports.jsonl", records)
        _write_json(output_dir / "fda_core_outbreak_table.json", core_meta)
        _write_json(
            output_dir / "fda_outbreaks_manifest.json",
            {
                "generated_at": datetime.now(UTC).isoformat(),
                "report_records": len(records),
                "reports_url": reports_url,
                "core_url": core_url,
            },
        )
        return len(records)

    def fetch_jecfa_records(self) -> int:
        output_dir = self.output_dir / "jecfa"
        output_dir.mkdir(parents=True, exist_ok=True)
        latest = _request_json("https://apps.who.int/food-additives-contaminants-jecfa-database/api/ChemicalData/GetLatest")
        results: dict[str, dict[str, Any]] = {}

        for item in latest:
            if any(term in (item.get("Name", "") or "").lower() for term in HAZARD_TERMS):
                results[str(item.get("Id"))] = item

        for term in HAZARD_TERMS:
            encoded = urllib.parse.quote(term)
            search_url = (
                "https://apps.who.int/food-additives-contaminants-jecfa-database/api/"
                f"SearchChemical/ByPartialName/{encoded}"
            )
            for item in _request_json(search_url):
                results[str(item.get("Id"))] = item
            time.sleep(0.1)

        rows = [
            {
                "record_type": "authority_evidence",
                "source": "jecfa",
                "source_url_or_id": f"https://apps.who.int/food-additives-contaminants-jecfa-database/Home/Chemical/{item.get('Id')}",
                "authority": "WHO/JECFA",
                "evidence_type": "toxicology",
                "name": item.get("Name", ""),
                "adi": item.get("ADI", ""),
                "cas_no": item.get("CAS_NO", ""),
                "fema_no": item.get("FEMA_NO", ""),
                "jecfa_no": item.get("JECFA_NO", ""),
                "functional_class": item.get("FunctionalClass", ""),
                "raw": item,
            }
            for item in results.values()
        ]

        _write_jsonl(output_dir / "jecfa_records.jsonl", rows)
        _write_json(
            output_dir / "jecfa_manifest.json",
            {
                "generated_at": datetime.now(UTC).isoformat(),
                "records": len(rows),
                "hazard_terms": HAZARD_TERMS,
            },
        )
        return len(rows)

    def fetch_openfoodtox_index(self) -> int:
        output_dir = self.output_dir / "openfoodtox"
        output_dir.mkdir(parents=True, exist_ok=True)
        source_url = "https://www.efsa.europa.eu/en/data-report/chemical-hazards-database-openfoodtox"
        html = _request_text(source_url)
        soup = BeautifulSoup(html, "html.parser")
        title = soup.title.string.strip() if soup.title and soup.title.string else ""
        description = ""
        desc = soup.find("meta", attrs={"name": "description"})
        if desc:
            description = (desc.get("content") or "").strip()
        links = []
        for anchor in soup.find_all("a", href=True):
            href = anchor["href"]
            if any(key in href.lower() for key in ["zenodo", "openfoodtox", "efsa", "doi.org/10.5281"]):
                links.append(href)
        payload = {
            "record_type": "authority_evidence",
            "source": "openfoodtox",
            "source_url_or_id": source_url,
            "authority": "EFSA",
            "evidence_type": "toxicology_index",
            "title": title,
            "description": description,
            "related_links": list(dict.fromkeys(links)),
        }
        _write_json(output_dir / "openfoodtox_index.json", payload)
        return len(payload["related_links"])


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Fetch public non-PubMed risk-taxonomy sources.")
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Output directory for public-source fetch results.",
    )
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    manifest = PublicRiskSourceFetcher(output_dir=Path(args.output_dir)).fetch_all()
    print(_json_dumps(manifest))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
